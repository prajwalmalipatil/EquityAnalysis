"""
formatters.py
Excel formatting and styling logic for VSA result files.
Uses openpyxl to apply conditional formatting and professional layouts.
"""

from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from typing import Optional, Dict

class ExcelFormatter:
    """
    Handles professional styling for generated Excel workbooks.
    Complies with modern UI standards requested by the user.
    """
    
    @staticmethod
    def apply_standard_styling(workbook, sheet_name: str) -> None:
        """Applies zebra striping and conditional coloring to a sheet."""
        try:
            if sheet_name not in workbook.sheetnames:
                return
                
            ws = workbook[sheet_name]
            max_row, max_col = ws.max_row, ws.max_column
            
            if max_row < 2:
                return
                
            headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
            
            def get_col_letter(name: str) -> Optional[str]:
                return get_column_letter(headers[name]) if name in headers else None

            # 1. Zebra Striping
            zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{max_row}",
                    FormulaRule(formula=["MOD(ROW(),2)=0"], fill=zebra_fill)
                )

            # 2. Signal Type Coloring
            signal_col = get_col_letter("Signal_Type")
            if signal_col:
                bullish_fill = PatternFill(start_color="B7E1CD", end_color="B7E1CD", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{signal_col}2:{signal_col}{max_row}",
                    FormulaRule(formula=[f'ISNUMBER(SEARCH("Bullish",{signal_col}2))'], fill=bullish_fill)
                )
                
                bearish_fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
                ws.conditional_formatting.add(
                    f"{signal_col}2:{signal_col}{max_row}",
                    FormulaRule(formula=[f'ISNUMBER(SEARCH("Bearish",{signal_col}2))'], fill=bearish_fill)
                )

            # 3. Auto-adjust columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        except Exception as e:
            # Minimal logging as this is a utility
            pass

    @staticmethod
    def add_vsa_legend(workbook) -> None:
        """Adds a legend sheet explaining the VSA signals."""
        if "Legend" in workbook.sheetnames:
            workbook.remove(workbook["Legend"])
        
        ws = workbook.create_sheet("Legend")
        ws.append(["Signal Category", "Description", "Sentiment"])
        
        legend_data = [
            ["Selling Climax", "Ultra high volume + wide spread + weak close in downtrend", "Bullish Reversal"],
            ["Buying Climax", "Ultra high volume + wide spread + strong close in uptrend", "Bearish Reversal"],
            ["No Demand", "Up bar on low volume - lack of buying interest", "Bearish Weakness"],
            ["Stopping Volume", "High volume + narrow spread + mid-close = absorption", "Potential Reversal"],
            ["Silent Accumulation", "High quality volume drop + close above previous", "Strong Bullish"],
        ]
        
        for row in legend_data:
            ws.append(row)
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
